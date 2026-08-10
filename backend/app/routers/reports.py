from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import User, UserRole
from ..security import get_current_user
from ..services import dashboard_overview

router = APIRouter(prefix="/reports", tags=["reports"])


def _data(db: Session, user: User):
    return dashboard_overview(db, team_id=user.team_id if user.role == UserRole.TEAM.value else None)


@router.get("/daily.xlsx")
def daily_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    data = _data(db, user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws.append(["GREEN HOLDING LIVE COMMAND CENTER"])
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    k = data["kpis"]
    ws.append(["GMV", "Orders", "Ads Spend", "AOV", "Ads/GMV %", "ROAS", "Net Revenue", "Refund %", "GMV/hour"])
    ws.append([k["gmv"], k["orders"], k["ads_spend"], k["aov"], k["ads_percentage"], k["roas"], k["net_revenue"], k["refund_rate"], k["gmv_per_hour"]])
    ws.append([])
    ws.append(["Ranking", "Team", "GMV", "Orders", "AOV", "Ads", "Ads/GMV %", "Refund %", "Net Revenue", "GMV/hour"])
    for idx, row in enumerate(data["ranking"], 1):
        ws.append([idx, row["team_name"], row["gmv"], row["orders"], row["aov"], row["ads"], row["ads_percentage"], row["refund_rate"], row["net"], row["gmv_per_hour"]])
    for idx, col in enumerate(ws.iter_cols(), 1):
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(28, max(12, width))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=green-holding-daily-report.xlsx"})


@router.get("/daily.pdf")
def daily_pdf(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    data = _data(db, user)
    buf = BytesIO()
    try:
        pdfmetrics.registerFont(TTFont("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
        font_name = "DejaVu"
    except Exception:
        font_name = "Helvetica"
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font_name
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    story = [Paragraph("GREEN HOLDING LIVE COMMAND CENTER — BÁO CÁO NGÀY", styles["Title"]), Spacer(1, 12)]
    k = data["kpis"]
    kpi_table = Table([
        ["GMV", "Orders", "Ads", "AOV", "Ads/GMV", "Net Revenue", "Refund"],
        [f"{k['gmv']:,.0f}", f"{k['orders']:,}", f"{k['ads_spend']:,.0f}", f"{k['aov']:,.0f}", f"{k['ads_percentage']:.2f}%", f"{k['net_revenue']:,.0f}", f"{k['refund_rate']:.2f}%"],
    ], repeatRows=1)
    kpi_table.setStyle(TableStyle([("FONTNAME", (0,0), (-1,-1), font_name), ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0b3d2a")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .4, colors.grey), ("ALIGN", (0,0), (-1,-1), "CENTER"), ("PADDING", (0,0), (-1,-1), 7)]))
    story += [kpi_table, Spacer(1, 16), Paragraph("Xếp hạng 4 team", styles["Heading2"])]
    rows = [["#", "Team", "GMV", "Orders", "AOV", "Ads", "Ads/GMV", "Refund", "Net"]]
    for idx, row in enumerate(data["ranking"], 1):
        rows.append([idx, row["team_name"], f"{row['gmv']:,.0f}", row["orders"], f"{row['aov']:,.0f}", f"{row['ads']:,.0f}", f"{row['ads_percentage']:.2f}%", f"{row['refund_rate']:.2f}%", f"{row['net']:,.0f}"])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([("FONTNAME", (0,0), (-1,-1), font_name), ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#112a20")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .4, colors.grey), ("PADDING", (0,0), (-1,-1), 6)]))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=green-holding-daily-report.pdf"})

# ---- Filterable report API -------------------------------------------------
from datetime import date, datetime, time, timedelta, timezone
from fastapi import Query
from sqlalchemy import func, select
from sqlalchemy.orm import joinedload

from ..models import LiveSession, Order
from ..services import as_float, serialize_session

VN_TZ = timezone(timedelta(hours=7))


def _range(date_from: date | None, date_to: date | None) -> tuple[datetime, datetime]:
    today = datetime.now(VN_TZ).date()
    start_date = date_from or today
    end_date = date_to or start_date
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    start = datetime.combine(start_date, time.min, VN_TZ).astimezone(timezone.utc)
    end = datetime.combine(end_date + timedelta(days=1), time.min, VN_TZ).astimezone(timezone.utc)
    return start, end


def report_summary(
    db: Session,
    user: User,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    team_id: int | None = None,
    channel_id: int | None = None,
    shift: str | None = None,
) -> dict:
    start, end = _range(date_from, date_to)
    query = select(LiveSession).options(joinedload(LiveSession.channel), joinedload(LiveSession.team)).where(
        LiveSession.started_at >= start, LiveSession.started_at < end
    )
    if user.role == UserRole.TEAM.value:
        query = query.where(LiveSession.team_id == user.team_id)
    elif team_id:
        query = query.where(LiveSession.team_id == team_id)
    if channel_id:
        query = query.where(LiveSession.channel_id == channel_id)
    if shift:
        query = query.where(LiveSession.shift == shift.upper())
    sessions = db.scalars(query.order_by(LiveSession.started_at.desc())).unique().all()
    rows = [serialize_session(db, s) for s in sessions]
    gmv = sum(x["gmv"] for x in rows); orders = sum(x["orders"] for x in rows); ads = sum(x["ads_spend"] for x in rows)
    refunds = sum(x["refund_amount"] + x["cancelled_amount"] for x in rows); net = max(0.0, gmv - refunds)
    duration_h = sum(x["duration_seconds"] for x in rows) / 3600
    rollup: dict[int, dict] = {}
    for x in rows:
        r = rollup.setdefault(x["team_id"], {"team_id":x["team_id"],"team_name":x["team_name"],"gmv":0.0,"orders":0,"ads":0.0,"net":0.0,"duration":0})
        r["gmv"] += x["gmv"]; r["orders"] += x["orders"]; r["ads"] += x["ads_spend"]; r["net"] += x["net_revenue"]; r["duration"] += x["duration_seconds"]
    ranking=[]
    for r in rollup.values():
        r["aov"] = r["gmv"] / r["orders"] if r["orders"] else 0
        r["ads_percentage"] = r["ads"] / r["gmv"] * 100 if r["gmv"] else 0
        r["refund_rate"] = (r["gmv"] - r["net"]) / r["gmv"] * 100 if r["gmv"] else 0
        r["gmv_per_hour"] = r["gmv"] / max(r["duration"] / 3600, 1/60)
        ranking.append(r)
    ranking.sort(key=lambda x:x["gmv"], reverse=True)
    session_ids=[s.id for s in sessions]
    top_skus=[]
    if session_ids:
        q=(select(Order.product_name,func.sum(Order.payment_amount),func.sum(Order.quantity)).where(Order.live_session_id.in_(session_ids)).group_by(Order.product_name).order_by(func.sum(Order.payment_amount).desc()).limit(10))
        top_skus=[{"name":r[0] or "Chưa có tên","revenue":as_float(r[1]),"quantity":int(r[2] or 0)} for r in db.execute(q).all()]
    return {
        "date_from": start.astimezone(VN_TZ).date(), "date_to": (end - timedelta(seconds=1)).astimezone(VN_TZ).date(),
        "filters":{"team_id":team_id,"channel_id":channel_id,"shift":shift},
        "kpis":{"gmv":gmv,"orders":orders,"ads_spend":ads,"aov":gmv/orders if orders else 0,"ads_percentage":ads/gmv*100 if gmv else 0,"roas":gmv/ads if ads else 0,"net_revenue":net,"refund_rate":refunds/gmv*100 if gmv else 0,"gmv_per_hour":gmv/duration_h if duration_h else 0},
        "ranking": ranking, "top_skus":top_skus, "sessions":rows,
    }


@router.get("/summary")
def summary(
    date_from: date | None = Query(default=None), date_to: date | None = Query(default=None),
    team_id: int | None = Query(default=None), channel_id: int | None = Query(default=None), shift: str | None = Query(default=None),
    db: Session = Depends(get_db), user: User = Depends(get_current_user),
):
    return report_summary(db,user,date_from=date_from,date_to=date_to,team_id=team_id,channel_id=channel_id,shift=shift)


def _xlsx_from_summary(data: dict) -> BytesIO:
    wb=Workbook(); ws=wb.active; ws.title="Report"
    ws.append(["GREEN HOLDING LIVE COMMAND CENTER"]); ws.merge_cells("A1:J1"); ws["A1"].font=Font(size=18,bold=True); ws["A1"].alignment=Alignment(horizontal="center")
    ws.append(["Từ ngày",str(data["date_from"]),"Đến ngày",str(data["date_to"])])
    k=data["kpis"]; ws.append(["GMV","Orders","Ads Spend","AOV","Ads/GMV %","ROAS","Net Revenue","Refund %","GMV/hour"]); ws.append([k["gmv"],k["orders"],k["ads_spend"],k["aov"],k["ads_percentage"],k["roas"],k["net_revenue"],k["refund_rate"],k["gmv_per_hour"]]); ws.append([])
    ws.append(["#","Team","GMV","Orders","AOV","Ads","Ads/GMV %","Refund %","Net Revenue","GMV/hour"])
    for i,r in enumerate(data["ranking"],1): ws.append([i,r["team_name"],r["gmv"],r["orders"],r["aov"],r["ads"],r["ads_percentage"],r["refund_rate"],r["net"],r["gmv_per_hour"]])
    ws.append([]);ws.append(["Session","Team","Kênh","Ca","Bắt đầu","Kết thúc","GMV","Orders","Ads","Net"])
    for r in data["sessions"]: ws.append([r["session_code"],r["team_name"],r["channel_name"],r["shift"],str(r["started_at"]),str(r["ended_at"] or ""),r["gmv"],r["orders"],r["ads_spend"],r["net_revenue"]])
    for idx,col in enumerate(ws.iter_cols(),1): ws.column_dimensions[get_column_letter(idx)].width=min(30,max(12,max(len(str(c.value or "")) for c in col)+2))
    buf=BytesIO();wb.save(buf);buf.seek(0);return buf


@router.get("/export.xlsx")
def export_xlsx(
    date_from: date | None = Query(default=None), date_to: date | None = Query(default=None), team_id: int | None = Query(default=None), channel_id: int | None = Query(default=None), shift: str | None = Query(default=None),
    db: Session = Depends(get_db), user: User = Depends(get_current_user),
):
    data=report_summary(db,user,date_from=date_from,date_to=date_to,team_id=team_id,channel_id=channel_id,shift=shift)
    return StreamingResponse(_xlsx_from_summary(data),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=green-holding-report.xlsx"})


@router.get("/export.pdf")
def export_pdf(
    date_from: date | None = Query(default=None), date_to: date | None = Query(default=None), team_id: int | None = Query(default=None), channel_id: int | None = Query(default=None), shift: str | None = Query(default=None),
    db: Session = Depends(get_db), user: User = Depends(get_current_user),
):
    data=report_summary(db,user,date_from=date_from,date_to=date_to,team_id=team_id,channel_id=channel_id,shift=shift);buf=BytesIO()
    try: pdfmetrics.registerFont(TTFont("DejaVu","/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"));font="DejaVu"
    except Exception: font="Helvetica"
    styles=getSampleStyleSheet();[setattr(x,"fontName",font) for x in styles.byName.values()]
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=24,rightMargin=24,topMargin=24,bottomMargin=24)
    story=[Paragraph(f"GREEN HOLDING LIVE COMMAND CENTER — {data['date_from']} → {data['date_to']}",styles["Title"]),Spacer(1,12)]
    k=data["kpis"]; kt=Table([["GMV","Orders","Ads","AOV","Ads/GMV","Net","Refund"],[f"{k['gmv']:,.0f}",f"{k['orders']:,}",f"{k['ads_spend']:,.0f}",f"{k['aov']:,.0f}",f"{k['ads_percentage']:.2f}%",f"{k['net_revenue']:,.0f}",f"{k['refund_rate']:.2f}%"]]);kt.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0b3d2a")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.grey),("PADDING",(0,0),(-1,-1),6)]));story += [kt,Spacer(1,14)]
    rows=[["#","Team","GMV","Orders","Ads","Refund","Net","GMV/h"]]+[[i,r["team_name"],f"{r['gmv']:,.0f}",r["orders"],f"{r['ads']:,.0f}",f"{r['refund_rate']:.2f}%",f"{r['net']:,.0f}",f"{r['gmv_per_hour']:,.0f}"] for i,r in enumerate(data["ranking"],1)]
    table=Table(rows,repeatRows=1);table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#112a20")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.grey),("PADDING",(0,0),(-1,-1),5)]));story.append(table);doc.build(story);buf.seek(0)
    return StreamingResponse(buf,media_type="application/pdf",headers={"Content-Disposition":"attachment; filename=green-holding-report.pdf"})
